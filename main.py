import tkinter as tk
from tkinter import *
import tkinter.font as tkFont
import random
from time import *
from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk


class MyProgram:
    window = tk.Tk()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    canvas = tk.Canvas(window, width=1200, height=1000)
    objects_for_study = 0 # 0 - картинки, 1 - слова, 2 - и картинки, и слова
    categoryA_train = '' # 1 слово
    categoryB_train = ''
    categoryC_train = ''
    categoryD_train = ''
    categoryA_main = ''
    categoryB_main = ''
    categoryC_main = ''
    categoryD_main = ''
    category1_now = ''
    category2_now = ''
    word_category = 0
    categories_now = 11 #11 - 1;1, 12-1;2, 21 - 2;1, 22-2;2
    train_firstA = 1 # 1 - сначала 1-ая категория, 2 - 2-ая
    main_firstA = 1 # 1 - сначала 1-ая категория, 2 - 2-ая
    categories_order = []
    errors_test = []
    wordsA_train = [] # набор слов
    wordsB_train = []
    wordsC_train = []
    wordsD_train = []
    wordsA_main = []
    wordsB_main = []
    wordsC_main = []
    wordsD_main = []
    order_block = []
    order_block_2 = []
    order_block_3 = []
    order_block_4 = []
    order_block_5 = []
    order_block_6 = []
    category_order_1 = []
    category_order_2 = []
    category_order_3 = []
    category_order_4 = []
    category_order_5 = []
    category_order_6 = []
    answers = []
    start_time = 0
    end_time = 0
    block1_speed = []
    block2_speed = []
    block3_speed = []
    block4_speed = []
    block5_speed = []
    block6_speed = []
    blocks_speed = []

    numb_for_train = 0
    numb_for_main = 0
    participant_number = 0
    participant_age = ''
    job = ""
    participant_work = 0
    participant_name = ''
    participant_gender = ''
    participant_fac = '-1'
    participant_cour = '-1'
    participant_form_of_edu = '-1'
    instr_name = "instructions\\"
    instr_numb = 0
    numb_of_block = 6
    block_right_now = 1
    counter = 0
    counterA = 0
    counterB = 0
    counterC = 0
    counterD = 0
    photos_name = "photos\\"
    test_name = "test_photos\\"
    main_object = ['Картинки', 'Слова', 'И картинки, и слова']
    loadExcel = False
    no_error = True
    train_rightnow = True
    error_img = 0
    train_random = 0
    main_random = 0
    instr_now = tk.IntVar()
    error_ei = tk.IntVar()
    genders = ['М', 'Ж']
    colors = ["Черный", "Красный", "Зеленый", "Синий", "Оранжевый", "Желтый", "Фиолетовый", "Розовый"]
    codes_of_colors = ['#000000', '#ff0000', '#008b00', '#0000ff', '#ffa500', '#c71585', '#5d478b', '#d74894']
    indexA = -1
    indexB = -1
    size_of_font = 10
    works = ['Студент', 'Школьник', 'Руководитель предприятия, учреждения', 'Инженерно-технический работник',
             'Служащий аппарата управления предприятия, учреждения',
             'Военнослужащий, работник правоохранительных органов, юстиции', 'Рабочий', 'Домохозяйка',
             'Предприниматель', 'Работник сферы обслуживания', 'Безработный, временно неработающий',
             'Представитель интеллигенции (образование, здравоохранение, культура)', 'Пенсионер',
             'Другой вид деятельности (укажите)']
    faculties = ['Авиационный техникум', 'Биологический факультет', 'Естественнонаучный институт', 'Институт авиационной и ракетно-космической техники', 'Институт двигателей и энергетических установок', 'Институт дополнительного образования', 'Институт информатики и кибернетики',
                 'Институт экономики и управления', 'Исторический факультет', 'Механико-математический факультет', 'Передовая инженерная аэрокосмическая школа', 'Психологический факультет', 'Социологический факультет', 'Факультет филологии и журналистики', 'Физический факультет',
                 'Химический факультет', 'Юридический институт']
    form_of_education = ['Бакалавриат', 'Магистратура', 'Специалитет', 'Аспирантура']
    bus_numb = ['1', '2', '3', '4', '5', '6', '7']
    bus_yes_no = ['Да', 'Нет']
    bus_job = ["Наемный работник", "Учредитель (предприниматель) в своей собственной фирме", "Преемник фирме родителей/семейной фирме или в другом бизнесе"]
    bus_ans = []

    def __init__(self):
        self.file1 = Workbook()
        self.window.title('ИАТ предприниматель')
        self.window.geometry(f'{self.screen_width}x{self.screen_height}')
        self.menubar = tk.Menu(self.window)
        self.window.config(menu=self.menubar)
        self.canvas.pack()
        self.instr = PhotoImage()
        self.warning = PhotoImage()
        self.instr_list = []
        self.instr_numb = 1
        self.warning_check = tk.IntVar()
        settings_menu = tk.Menu(self.menubar, tearoff=0)
        settings_menu.add_command(label='Старт', command=self.start_test)
        # settings_menu.add_command(label='Изображения', command=self.create_settings_win)
        self.menubar.add_cascade(label='Настройки', menu=settings_menu)
        # image1 = Image.open("BackgrBack.png")
        # test = ImageTk.PhotoImage(image1)
        # label1 = tk.Label(image=test)
        # label1.image = test
        # label1.place(x=1,y=1)
        # image2 = Image.open("1B.png")
        # test1 = ImageTk.PhotoImage(image2)
        # label2 = tk.Label(image=test1)
        # label2.image = test1
        # label2.place(x=2, y=2)


        # self.window.attributes('-fullscreen', True)

    @staticmethod
    def start():
        MyProgram.window.mainloop()

    def start_test(self):
        self.ask_about_excel()
        self.categoryA_main = 'Предприниматель'
        self.categoryB_main = 'Наемный работник'
        self.categoryC_main = 'Хорошо'
        self.categoryD_main = 'Плохо'
        self.categoryA_train = 'Цветы'
        self.categoryB_train = 'Насекомое'
        self.categoryC_train = 'Приятные слова'
        self.categoryD_train = 'Неприятные слова'
        self.wordsA_train = ['Тюльпан', 'Роза', 'Нарцисс', 'Сирень', 'Лилия']
        self.wordsB_train = ['Оса', 'Блоха', 'Моль', 'Клоп', 'Комар']
        self.wordsC_train = ['Радость', 'Наслаждение', 'Счастье', 'Блаженство', 'Мир']
        self.wordsD_train = ['Отвращение', 'Страдание', 'Яд', 'Зло', 'Боль']
        self.wordsA_main = ['Предприниматель', 'Бизнесмен', 'Собственник фирмы', 'Владелец бизнеса', 'Самозанятый']
        self.wordsB_main = ['Работник', 'Сотрудник', 'Подчинённый', 'Трудящийся', 'Рабочий']
        self.wordsC_main = ['Любовь', 'Успех', 'Довольный', 'Веселье', 'Победа']
        self.wordsD_main = ['Гнилой', 'горе', 'Предательство', 'Болезнь', 'Смерть']
        # self.categoryA_main = '1'
        # self.categoryB_main = '2'
        # self.categoryC_main = '3'
        # self.categoryD_main = '4'
        # self.categoryA_train = '5'
        # self.categoryB_train = '6'
        # self.categoryC_train = '7'
        # self.categoryD_train = '8'
        # self.wordsA_main = ['11']
        # self.wordsB_main = ['22']
        # self.wordsC_main = ['33']
        # self.wordsD_main = ['44']
        # self.wordsA_train = ['55']
        # self.wordsB_train = ['66']
        # self.wordsC_train = ['77']
        # self.wordsD_train = ['88']
        if (len(self.wordsA_main) == 0 or len(self.wordsB_main) == 0 or len(self.wordsC_main) == 0 or len(
                self.wordsD_main) == 0 or
                len(self.wordsA_train) == 0 or len(self.wordsB_train) == 0 or len(self.wordsC_train) == 0 or len(
                    self.wordsD_train) == 0):
            check = tk.IntVar()
            win_settings = tk.Toplevel(self.window)
            win_settings.wm_title('Ошибка')
            win_settings.resizable(False, False)
            app_width = 225
            app_height = 125
            x = (self.screen_width / 2) - (app_width / 2)
            y = (self.screen_height / 2) - (app_height / 2)
            win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
            tk.Label(win_settings, text='Настройки не полностью заполнены').grid(row=0, column=1)
            tk.Button(win_settings, text='Ок', command=lambda: check.set(1)).grid(row=1, column=1)
            win_settings.wait_variable(check)
            win_settings.destroy()
        else:
            self.window.wm_attributes('-transparentcolor', 'cyan')
            self.show_warning()
            self.window.wait_variable(self.warning_check)
            self.get_subject_info()
            # self.load_instr()
            # self.show_instr()
            self.train_firstA = random.randint(1, 2)
            self.main_firstA = random.randint(1, 2)
            self.order_block_1 = random.sample(range(0, self.numb_for_train), self.numb_for_train)
            self.order_block_2 = random.sample(range(0, self.numb_for_train), self.numb_for_train)
            self.order_block_3 = random.sample(range(0, self.numb_for_main), self.numb_for_main)
            self.order_block_4 = random.sample(range(0, self.numb_for_main), self.numb_for_main)
            self.order_block_5 = random.sample(range(0, self.numb_for_main), self.numb_for_main)
            self.order_block_6 = random.sample(range(0, self.numb_for_main), self.numb_for_main)
            self.train_random = random.randint(0, 3)
            self.main_random = random.randint(0, 3)
            self.show_blocks()

    def ask_about_excel(self):
        check = tk.IntVar()
        var = tk.BooleanVar()
        var.set(True)
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Excel')
        win_settings.resizable(False, False)
        app_width = 250
        app_height = 150
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        # tk.Label(win_settings, text='Открыть существующий Excel?').grid(row=0, column=0)
        tk.Checkbutton(win_settings, text='Открыть существующий Excel?', variable=var, onvalue=1,
                       offvalue=0, font=("Arial", 10)).grid(row=1, column=0)
        tk.Button(win_settings, text='Ок', command=lambda: check.set(1), font=("Arial", 10)).grid(row=2, column=0)
        win_settings.wait_variable(check)
        self.loadExcel = var.get()
        if self.loadExcel:
            self.ask_participant_number()
        win_settings.destroy()

    def show_warning(self):
        self.win_test = tk.Toplevel(self.window)
        self.win_test.wm_title('Информированное согласие')
        self.win_test.resizable(False, False)
        self.win_test.configure(bg='white')
        app_width = 850
        app_height = 600
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        self.win_test.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        # try:
        #     self.warning = PhotoImage(file=self.instr_name + str(1) + ".png")
        #     self.window.bind("<KeyPress-space>", self.delete_warning)
        #     self.canvas.delete("all")
        #     self.canvas.create_image(20, 20, anchor="nw", image=self.warning)
        # except TclError:
        #     self.error(5)
        info_label = tk.Label(self.win_test,
                              text='Информированное согласие',
                              font=("Arial", 20, 'bold'), bg='white')
        info_label.place(relx=0.25)
        info_label.configure(bg='white')
        category_label = tk.Label(self.win_test, text='Далее вам будет предложено пройти тест неявных ассоциаций, в котором вы будете \n как можно быстрее сортировать слова по категориям. ', font=("Arial", 14), bg='white')
        category_label.place(relx=0.05, rely=0.1)
        category_label.configure(bg='white')
        label_3 = tk.Label(self.win_test, text= 'В этом тесте измеряются отношения, предпочтения и убеждения, связанные с работой', font=("Arial", 14), bg='white')
        label_3.place(relx=0.05, rely=0.2)
        label_3.configure(bg='white')
        label_4 = tk.Label(self.win_test, text= 'Условия участия в исследовании Вы можете принять участие в исследовании, если: ', font=("Arial", 14), bg='white')
        label_4.place(relx=0.05, rely=0.25)
        label_4.configure(bg='white')
        label_5 = tk.Label(self.win_test, text='- Вы являетесь дееспособным (не страдающим тяжкими психическими расстройствами и \n расстройствами, способными ухудшить Ваше состояние или спровоцировать обострение) \n'
                                               '- У Вас нормальное или скорректированное зрение; \n - Вы находитесь в нормальном функциональном состоянии; \n - Вы согласны на обработку Ваших персональных данных без их раскрытия третьим \n лицам',
                           font=("Arial", 14), bg='white')
        label_5.place(relx=0.05, rely=0.3)
        label_5.configure(bg='white')
        label_6 = tk.Label(self.win_test, text='Участие в этом исследовании является добровольным, и вы можете прекратить \n выполнять задание в любое время.',
                           font=("Arial", 14), bg='white')
        label_6.place(relx=0.05, rely=0.55)
        label_6.configure(bg='white')
        label_7 = tk.Label(self.win_test, text='Все данные, собранные в исследовании, являются строго конфиденциальными. ', font=("Arial", 14, 'bold'), bg='white')
        label_7.place(relx=0.05, rely=0.65)
        label_7.configure(bg='white')
        label_8 = tk.Label(self.win_test, text='Нажимая пробел и переходя к выполнению заданий, вы соглашаетесь с участием в \n исследовании', font=("Arial", 14), bg='white')
        label_8.place(relx=0.05, rely=0.8)
        label_8.configure(bg='white')
        self.win_test.bind("<KeyPress-space>", self.delete_warning)
        self.window.bind("<KeyPress-space>", self.delete_warning)

    def delete_warning(self, event):
        # self.canvas.delete("all")
        self.win_test.unbind("<KeyPress-space>")
        self.win_test.destroy()
        self.window.unbind("<KeyPress-space>")

        self.warning_check.set(1)

    def show_instr(self):
        self.window.bind("<KeyPress-space>", self.next_instr)
        self.canvas.delete("all")
        if not self.objects_for_study: #pics
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[1], tags="first")
        elif self.objects_for_study == 1: #words:
            self.canvas.create_image(20, 20, anchor="nw", image=self.instr_list[1], tags="first")

    def next_instr(self, event):
        self.window.unbind("<KeyPress-space>")
        self.canvas.delete("first")

    def load_instr(self):
        try:
            self.instr = PhotoImage(file=self.instr_name + str(1) + ".png")
        except TclError:
            self.error(4)

    def show_example(self):
        self.win_test = tk.Toplevel(self.window)
        self.win_test.wm_title('Тест')
        self.win_test.resizable(False, False)
        self.win_test.configure(bg='white')
        app_width = 800
        app_height = 450
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        self.win_test.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        self.indexA = 3
        self.indexB = 5
        categoryA = ''
        categoryB = ''
        categoryC = ''
        categoryD = ''
        wordsA = []
        wordsB = []
        wordsC = []
        wordsD = []
        if self.train_rightnow:  # если тренировка (блок 1 и 2)
            categoryA = str(self.categoryA_train)
            categoryB = str(self.categoryB_train)
            categoryC = str(self.categoryC_train)
            categoryD = str(self.categoryD_train)
            for i in range(4):
                if (i >= len(self.wordsA_train)):
                    break
                else:
                    wordsA.append(self.wordsA_train[i])
                    wordsB.append(self.wordsB_train[i])
                    wordsC.append(self.wordsC_train[i])
                    wordsD.append(self.wordsD_train[i])

        else:
            categoryA = str(self.categoryA_main)
            categoryB = str(self.categoryB_main)
            categoryC = str(self.categoryC_main)
            categoryD = str(self.categoryD_main)
            for i in range(4):
                if (i >= len(self.wordsA_main)):
                    break
                else:
                    wordsA.append(self.wordsA_main[i])
                    wordsB.append(self.wordsB_main[i])
                    wordsC.append(self.wordsC_main[i])
                    wordsD.append(self.wordsD_main[i])
        step = 0.05
        move_x = 0.25
        move_y = 0.18
        info_label = tk.Label(self.win_test, text = ' Далее вы будете использовать компьютерные клавиши «E» и «I», \n чтобы как можно быстрее сортировать примеры по группам. \n Ознакомитесь с четырьмя группами категорий и их примерами:', font=("Arial", 15), bg='white')
        info_label.place(anchor="nw")
        info_label.configure(bg='white')
        category_label = tk.Label(self.win_test, text='Категория', font=("Arial", 12, 'bold'), bg='white')
        category_label.place(rely=move_y)
        category_label.configure(bg='white')
        example_label = tk.Label(self.win_test, text='Пример', font=("Arial", 12, 'bold'), bg='white')
        example_label.place(relx=move_x, rely=move_y)
        example_label.configure(bg='white')
        name_category1_label = tk.Label(self.win_test, text=str(categoryA), font=("Arial", 12), bg='white')
        name_category1_label.place(rely=move_y + step * 1)
        name_category1_label.configure(bg='white')
        name_category2_label = tk.Label(self.win_test, text=str(categoryB), font=("Arial", 12), bg='white')
        name_category2_label.place(rely=move_y + step * 2)
        name_category2_label.configure(bg='white')
        name_category3_label = tk.Label(self.win_test, text=str(categoryC), font=("Arial", 12), bg='white')
        name_category3_label.place(rely=move_y + step * 3)
        name_category3_label.configure(bg='white')
        name_category4_label = tk.Label(self.win_test, text=str(categoryD), font=("Arial", 12), bg='white')
        name_category4_label.place(rely=move_y + step * 4)
        name_category4_label.configure(bg='white')
        name_example1_label = tk.Label(self.win_test, text=str((','.join(wordsA))), font=("Arial", 12), bg='white')
        name_example1_label.place(relx=move_x, rely=move_y + step * 1)
        name_example1_label.configure(bg='white')
        name_example2_label = tk.Label(self.win_test, text=str((','.join(wordsB))), font=("Arial", 12), bg='white')
        name_example2_label.place(relx=move_x, rely=move_y + step * 2)
        name_example2_label.configure(bg='white')
        name_example3_label = tk.Label(self.win_test, text=str((','.join(wordsC))), font=("Arial", 12), bg='white')
        name_example3_label.place(relx=move_x, rely=move_y + step * 3)
        name_example3_label.configure(bg='white')
        name_example4_label = tk.Label(self.win_test, text=str((','.join(wordsD))), font=("Arial", 12), bg='white')
        name_example4_label.place(relx=move_x, rely=move_y + step * 4)
        name_example4_label.configure(bg='white')
        exit_label = tk.Label(self.win_test, text='Нажмите пробел, чтобы продолжить', font=("Arial", 15), bg='white')
        exit_label.place(relx=0.25, rely=0.75)
        exit_label.configure(bg='white')
        self.win_test.bind("<KeyPress-space>", self.destroy_example_window)
        self.window.bind("<KeyPress-space>", self.destroy_example_window)

    def destroy_example_window(self, event):
        self.win_test.unbind("<KeyPress-space>")
        self.window.unbind("<KeyPress-space>")
        self.win_test.destroy()
        self.instr_now.set(1)

    def show_blocks(self):
        self.numb_for_train = len(self.wordsA_train)
        self.numb_for_main = len(self.wordsA_main)
        self.show_example()
        for self.block_right_now in range(self.numb_of_block):
            self.counterA = 0
            self.counterB = 0
            self.counterC = 0
            self.counterD = 0
            self.counter = 0
            if self.block_right_now <= 1:
                self.order_block = random.sample(range(0, self.numb_for_train), self.numb_for_train)
            else:
                self.order_block = random.sample(range(0, self.numb_for_main), self.numb_for_main)
            self.make_category_order()
            if not self.block_right_now:
                self.win_test.wait_variable(self.instr_now)
            self.show_instr_window()
            self.win_test.wait_variable(self.instr_now)
            if self.block_right_now <= 1:
                for self.counter in range(4 * len(self.wordsA_train)):
                    self.window.after(400, self.show_test_window())
                    self.win_test.wait_variable(self.instr_now)
            else:
                for self.counter in range(4 * len(self.wordsA_main)):
                    self.window.after(400, self.show_test_window())
                    self.win_test.wait_variable(self.instr_now)
        self.test_end()

    def test_end(self):
        self.ask_bussines_questions()
        self.create_excel()
        check = tk.IntVar()
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Тест окончен')
        win_settings.resizable(False, False)
        app_width = 230
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Тестирование успешно завершено!').grid(row=0, column=1)
        tk.Button(win_settings, text='Ок', command=lambda: check.set(1)).grid(row=1, column=1)
        win_settings.wait_variable(check)
        self.clear_lists()
        win_settings.destroy()

    def ask_bussines_questions(self):
        check = tk.IntVar()
        # self.participant_work = self.works[0]
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Доп вопросы')
        win_settings.resizable(False, False)
        app_width = 1300
        app_height = 750
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        font1 = tkFont.Font(family='Arial', size=12)
        tk.Label(win_settings, text='Вы имеете собственное дело, бизнес (независимый предприниматель,'
                                    ' работаете на собственном предприятии, ферме, занимаетесь индивидуальной трудовой деятельностью и т. п.)? ', font=("Arial", 11) ).grid(row=0, column=0)
        tk.Label(win_settings, text='Являетесь ли вы в настоящее время — самостоятельно или вместе с другими — владельцем какого-либо бизнеса? ', font=("Arial", 11)).grid(row=2, column=0)
        yes_no_Obj_1 = tk.StringVar(win_settings)
        yes_no_Obj_1.set(self.bus_yes_no[0])
        yes_no_Menu1 = tk.OptionMenu(win_settings, yes_no_Obj_1, *self.bus_yes_no)
        yes_no_Menu1.grid(row=1, column=0)
        yes_no_Menu1.configure(font=font1)
        yes_no_Obj_2 = tk.StringVar(win_settings)
        yes_no_Obj_2.set(self.bus_yes_no[0])
        yes_no_Menu2 = tk.OptionMenu(win_settings, yes_no_Obj_2, *self.bus_yes_no)
        yes_no_Menu2.grid(row=3, column=0)
        yes_no_Menu2.configure(font=font1)
        job_Obj_1 = tk.StringVar(win_settings)
        job_Obj_2 = tk.StringVar(win_settings)
        if self.participant_work == 0 or self.participant_work == 1:
            tk.Label(win_settings, text='Какую карьеру вы выберете сразу после окончания учебы? ', font=("Arial", 11)).grid(row=4, column=0)
            tk.Label(win_settings, text='Какую карьеру вы выберете сразу через 5 лет после окончания учебы?', font=("Arial", 11)).grid(row=6, column=0)
            job_Obj_1.set(self.bus_job[0])
            job_Menu1 = tk.OptionMenu(win_settings, job_Obj_1, *self.bus_job)
            job_Menu1.grid(row=5, column=0)
            job_Menu1.configure(font=font1)
            job_Obj_2.set(self.bus_job[0])
            job_Menu2 = tk.OptionMenu(win_settings, job_Obj_2, *self.bus_job)
            job_Menu2.grid(row=7, column=0)
            job_Menu2.configure(font=font1)
        tk.Label(win_settings, text='Оцените пожалуйста, степень согласия с этими высказываниями по 7-балльной шкале: от 1— полностью не согласен, до 7 — полностью согласен', font=("Arial", 11)).grid(row=8, column=0)
        tk.Label(win_settings, text='Среди всех возможных вариантов, я бы предпочел(ла) стать предпринимателем ', font=("Arial", 11)).grid(row=9, column=0)
        tk.Label(win_settings, text='Работа в качестве предпринимателя принесла бы мне большее чувство удовлетворения', font=("Arial", 11)).grid(row=11, column=0)
        tk.Label(win_settings, text='Если бы у меня были возможности и ресурсы, я бы стал(а) предпринимателем', font=("Arial", 11)).grid(row=13, column=0)
        tk.Label(win_settings, text='Работа в качестве предпринимателя, в моем понимании, несет в себе больше преимуществ, чем недостатков ', font=("Arial", 11)).grid(row=15, column=0)
        tk.Label(win_settings, text='Я готов(а) предпринять все необходимые усилия для начала своего дела и управления своей собственной фирмой', font=("Arial", 11)).grid(row=17, column=0)
        tk.Label(win_settings, text='Я решительно настроен(а) на создание своей фирмы в будущем', font=("Arial", 11)).grid(row=19, column=0)
        tk.Label(win_settings, text='Я серьезно обдумываю возможность начала своего бизнеса ', font=("Arial", 11)).grid(row=21, column=0)
        numb_Obj_1 = tk.StringVar(win_settings)
        numb_Obj_1.set(self.bus_numb[0])
        numb_Menu1 = tk.OptionMenu(win_settings, numb_Obj_1, *self.bus_numb)
        numb_Menu1.grid(row=10, column=0)
        numb_Menu1.configure(font=font1)
        numb_Obj_2 = tk.StringVar(win_settings)
        numb_Obj_2.set(self.bus_numb[0])
        numb_Menu2 = tk.OptionMenu(win_settings, numb_Obj_2, *self.bus_numb)
        numb_Menu2.grid(row=12, column=0)
        numb_Menu2.configure(font=font1)
        numb_Obj_3 = tk.StringVar(win_settings)
        numb_Obj_3.set(self.bus_numb[0])
        numb_Menu3 = tk.OptionMenu(win_settings, numb_Obj_3, *self.bus_numb)
        numb_Menu3.grid(row=14, column=0)
        numb_Menu3.configure(font=font1)
        numb_Obj_4 = tk.StringVar(win_settings)
        numb_Obj_4.set(self.bus_numb[0])
        numb_Menu4 = tk.OptionMenu(win_settings, numb_Obj_4, *self.bus_numb)
        numb_Menu4.grid(row=16, column=0)
        numb_Menu4.configure(font=font1)
        numb_Obj_5 = tk.StringVar(win_settings)
        numb_Obj_5.set(self.bus_numb[0])
        numb_Menu5 = tk.OptionMenu(win_settings, numb_Obj_5, *self.bus_numb)
        numb_Menu5.grid(row=18, column=0)
        numb_Menu5.configure(font=font1)
        numb_Obj_6 = tk.StringVar(win_settings)
        numb_Obj_6.set(self.bus_numb[0])
        numb_Menu6 = tk.OptionMenu(win_settings, numb_Obj_6, *self.bus_numb)
        numb_Menu6.grid(row=20, column=0)
        numb_Menu6.configure(font=font1)
        numb_Obj_7 = tk.StringVar(win_settings)
        numb_Obj_7.set(self.bus_numb[0])
        numb_Menu7 = tk.OptionMenu(win_settings, numb_Obj_7, *self.bus_numb)
        numb_Menu7.grid(row=22, column=0)
        numb_Menu7.configure(font=font1)
        tk.Button(win_settings, text='Применить', command=lambda: check.set(1), font=("Arial", 12)).grid(row=23, column=0)
        win_settings.wait_variable(check)
        self.bus_ans.append(yes_no_Obj_1.get())
        self.bus_ans.append(yes_no_Obj_2.get())
        self.bus_ans.append(job_Obj_1.get())
        self.bus_ans.append(job_Obj_2.get())
        self.bus_ans.append(numb_Obj_1.get())
        self.bus_ans.append(numb_Obj_2.get())
        self.bus_ans.append(numb_Obj_3.get())
        self.bus_ans.append(numb_Obj_4.get())
        self.bus_ans.append(numb_Obj_5.get())
        self.bus_ans.append(numb_Obj_6.get())
        self.bus_ans.append(numb_Obj_7.get())
        win_settings.destroy()
        self.change_bus_ans()

    def change_bus_ans(self):
        temp = []
        for i in range(2):
            if self.bus_ans[i] == 'Да':
                temp.append(1)
            else:
                temp.append(2)
        if not self.participant_work or self.participant_work == 1:
            for i in range(2, 4):
                if self.bus_ans[i] == 'Наемный работник':
                    temp.append(1)
                elif self.bus_ans[i] == 'Учредитель (предприниматель) в своей собственной фирме':
                    temp.append(2)
                else:
                    temp.append(3)
        else:
            for i in range(2, 4):
                temp.append(0)
        for i in range(4):
            self.bus_ans[i] = int(temp[i])

    def make_category_order(self):
        self.instr_now.set(0)
        temp = []
        if self.block_right_now <= 1:
            for i in range(len(self.wordsA_train)):
                temp.append(3)
                temp.append(2)
                temp.append(1)
                temp.append(0)
            self.category_order_1 = random.sample(temp, len(temp))
            self.category_order_2 = random.sample(temp, len(temp))
        else:
            for i in range(len(self.wordsA_main)):
                temp.append(3)
                temp.append(2)
                temp.append(1)
                temp.append(0)
            self.category_order_3 = random.sample(temp, len(temp))
            self.category_order_4 = random.sample(temp, len(temp))
            self.category_order_5 = random.sample(temp, len(temp))
            self.category_order_6 = random.sample(temp, len(temp))
        self.instr_now.set(1)

    def show_instr_window(self):
        self.instr_now.set(0)
        self.win_test = tk.Toplevel(self.window)
        self.win_test.wm_title('Тест')
        self.win_test.resizable(False, False)
        self.win_test.configure(bg='white')
        app_width = 800
        app_height = 550
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        self.win_test.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        words1 = []
        words2 = []
        if self.block_right_now <= 1:
            if self.train_firstA == 1:
                if not self.block_right_now:
                    self.category1_now = self.categoryA_train
                    self.category2_now = self.categoryC_train
                    self.categories_now = 11
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_train)):
                            break
                        else:
                            words1.append(self.wordsA_train[i])
                            words2.append(self.wordsC_train[i])
                else:
                    self.category1_now = self.categoryB_train
                    self.category2_now = self.categoryC_train
                    self.categories_now = 21
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_train)):
                            break
                        else:
                            words1.append(self.wordsB_train[i])
                            words2.append(self.wordsC_train[i])
            else:
                if not self.block_right_now:
                    self.category1_now = self.categoryB_train
                    self.category2_now = self.categoryC_train
                    self.categories_now = 21
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_train)):
                            break
                        else:
                            words1.append(self.wordsB_train[i])
                            words2.append(self.wordsC_train[i])
                else:
                    self.category1_now = self.categoryA_train
                    self.category2_now = self.categoryC_train
                    self.categories_now = 11
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_train)):
                            break
                        else:
                            words1.append(self.wordsA_train[i])
                            words2.append(self.wordsC_train[i])
        else:
            if self.main_firstA == 1:
                if self.block_right_now == 2 or self.block_right_now == 4:
                    self.category1_now = self.categoryA_main
                    self.category2_now = self.categoryC_main
                    self.categories_now = 11
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_main)):
                            break
                        else:
                            words1.append(self.wordsA_main[i])
                            words2.append(self.wordsC_main[i])
                else:
                    self.category1_now = self.categoryB_main
                    self.category2_now = self.categoryC_main
                    self.categories_now = 21
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_main)):
                            break
                        else:
                            words1.append(self.wordsB_main[i])
                            words2.append(self.wordsC_main[i])
            else:
                if self.block_right_now == 2 or self.block_right_now == 4:
                    self.category1_now = self.categoryB_main
                    self.category2_now = self.categoryC_main
                    self.categories_now = 21
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_main)):
                            break
                        else:
                            words1.append(self.wordsB_main[i])
                            words2.append(self.wordsC_main[i])
                else:
                    self.category1_now = self.categoryA_main
                    self.category2_now = self.categoryC_main
                    self.categories_now = 11
                    self.categories_order.append(self.categories_now)
                    for i in range(4):
                        if (i >= len(self.wordsA_main)):
                            break
                        else:
                            words1.append(self.wordsA_main[i])
                            words2.append(self.wordsC_main[i])

        e_but_label = tk.Label(self.win_test, text='"E" для всего остального', font=("Arial", 11), bg='white')
        e_but_label.place(relx=0.05, rely=0.02)
        i_but_label = tk.Label(self.win_test, text='"I" если принадлежит', font=("Arial", 11), bg='white')
        i_but_label.place(relx=0.8, rely=0.02)
        category1_label = tk.Label(self.win_test, text=self.category1_now, font=("Arial", 15), bg='white', fg=str(self.codes_of_colors[self.indexA]))
        category1_label.place(relx=0.5, rely=0.06)
        category1_words_label = tk.Label(self.win_test, text=str((','.join(words1))), font=("Arial", 13), bg='white', fg=str(self.codes_of_colors[self.indexA]))
        category1_words_label.place(relx=0.4, rely=0.12)
        and_label = tk.Label(self.win_test, text='и', font=("Arial", 13), bg='white')
        and_label.place(relx=0.5, rely=0.24)
        category2_label = tk.Label(self.win_test, text=self.category2_now, font=("Arial", 15), bg='white',
                                   fg=str(self.codes_of_colors[self.indexB]))
        category2_label.place(relx=0.5, rely=0.32)
        category2_words_label = tk.Label(self.win_test, text=str((','.join(words2))), font=("Arial", 13), bg='white',
                                         fg=str(self.codes_of_colors[self.indexB]))
        category2_words_label.place(relx=0.4, rely=0.38)
        block_number_label = tk.Label(self.win_test,
                                      text='Часть ' + str(self.block_right_now + 1) + ' из ' + str(self.numb_of_block),
                                      bg='white', font=("Arial", 12))
        block_number_label.place(relx=0.5, rely=0.5)
        underline_font = tkFont.Font(block_number_label, block_number_label.cget("font"))
        underline_font.configure(underline=True)
        block_number_label.configure(font=underline_font)
        info_i_label = tk.Label(self.win_test, text='Приготовьтесь нажимать указательным пальцем правой руки на клавишу "I" при появлении слов, \n которые относятся к категории  "' + str(self.category1_now) + '" ,  и к категории "' + str(self.category2_now) + '"'
                              , bg='white', font=("Arial", 13))
        info_i_label.place(rely=0.60)
        info_e_label = tk.Label(self.win_test, text='Приготовьтесь нажимать указательным пальцем левой руки на клавишу "E" \n при появлении слов, которые не относятся к этим категориям. '
                                , bg='white', font=("Arial", 13))
        info_e_label.place(rely=0.72)
        info_error_label = tk.Label(self.win_test, text='Если вы допустите ошибку появится красный крестик. Нажмите другую клавишу, чтобы продолжить'
                                    , bg='white', font=("Arial", 13))
        info_error_label.place(rely=0.85)
        start_label = tk.Label(self.win_test, text='Нажмите клавишу пробела, когда будете готовы к запуску'
                               , bg='white', font=("Arial", 13))
        start_label.place(relx=0.25, rely=0.95)
        self.win_test.bind("<KeyPress-space>", self.destroy_example_window)
        self.window.bind("<KeyPress-space>", self.destroy_example_window)

    def show_test_window(self):
        self.instr_now.set(0)
        self.error_ei.set(0)
        self.win_test = tk.Toplevel(self.window)
        self.win_test.wm_title('Тест')
        self.win_test.resizable(False, False)
        self.win_test.configure(bg='white')
        app_width = 730
        app_height = 550
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        self.win_test.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        word = ''
        color = ''
        category_order = []
        if not self.block_right_now:
            category_order = self.category_order_1
        elif self.block_right_now == 1:
            category_order = self.category_order_2
        elif self.block_right_now == 2:
            category_order = self.category_order_3
        elif self.block_right_now == 3:
            category_order = self.category_order_4
        elif self.block_right_now == 4:
            category_order = self.category_order_5
        else:
            category_order = self.category_order_6
        if self.block_right_now <= 1:
            if not category_order[self.counter]:
                word = self.wordsA_train[self.order_block[self.counterA]]
                self.counterA +=1
                self.word_category = 0
                color = str(self.codes_of_colors[self.indexA])
            elif category_order[self.counter] == 1:
                word = self.wordsB_train[self.order_block[self.counterB]]
                self.counterB += 1
                self.word_category = 1
                color = str(self.codes_of_colors[self.indexA])
            elif category_order[self.counter] == 2:
                word = self.wordsC_train[self.order_block[self.counterC]]
                self.counterC += 1
                self.word_category = 2
                color = str(self.codes_of_colors[self.indexB])
            else:
                word = self.wordsD_train[self.order_block[self.counterD]]
                self.counterD += 1
                self.word_category = 3
                color = str(self.codes_of_colors[self.indexB])
        else:
            if not category_order[self.counter]:
                word = self.wordsA_main[self.order_block[self.counterA]]
                self.counterA += 1
                self.word_category = 0
                color = str(self.codes_of_colors[self.indexA])
            elif category_order[self.counter] == 1:
                word = self.wordsB_main[self.order_block[self.counterB]]
                self.counterB += 1
                self.word_category = 1
                color = str(self.codes_of_colors[self.indexA])
            elif category_order[self.counter] == 2:
                word = self.wordsC_main[self.order_block[self.counterC]]
                self.counterC += 1
                self.word_category = 2
                color = str(self.codes_of_colors[self.indexB])
            else:
                word = self.wordsD_main[self.order_block[self.counterD]]
                self.counterD += 1
                self.word_category = 3
                color = str(self.codes_of_colors[self.indexB])

        e_but_label = tk.Label(self.win_test, text='"E" для всего остального', font=("Arial", 12), bg='white')
        e_but_label.place(relx=0.05, rely=0.02)
        i_but_label = tk.Label(self.win_test, text='"I" если принадлежит', font=("Arial", 12), bg='white')
        i_but_label.place(relx=0.75, rely=0.02)
        category1_label = tk.Label(self.win_test, text=self.category1_now, font=("Arial", 20), bg='white',
                                   fg=str(self.codes_of_colors[self.indexA]))
        category1_label.place(relx=0.45, rely=0.13)
        and_label = tk.Label(self.win_test, text='и', font=("Arial", 14), bg='white')
        and_label.place(relx=0.45, rely=0.19)
        category2_label = tk.Label(self.win_test, text=self.category2_now, font=("Arial", 20), bg='white',
                                   fg=str(self.codes_of_colors[self.indexB]))
        category2_label.place(relx=0.45, rely=0.25)
        word_label = tk.Label(self.win_test, text=word, font=("Arial", 20), bg='white', fg=color)
        word_label.place(relx=0.45, rely=0.5)
        # error_info_label = tk.Label(self.win_test, text='Если вы допустите ошибку, появится красный крестик. Нажмите другую клавишу, \n чтобы продолжить'
        #                       , font=("Arial", 10), bg='white')
        # error_info_label.place(relx=0.25, rely=0.8)
        self.win_test.bind("<KeyPress-e>", self.e_press)
        self.win_test.bind("<KeyPress-i>", self.i_press)
        self.window.bind("<KeyPress-e>", self.e_press)
        self.window.bind("<KeyPress-i>", self.i_press)
        self.start_time = time()

    def e_press(self, event):
        self.end_time = time()
        self.blocks_speed.append(self.end_time - self.start_time)
        self.answers.append('e')
        self.win_test.unbind("<KeyPress-e>")
        self.window.unbind("<KeyPress-e>")
        right = self.test_error('e')
        if not right:
            self.error_img = ImageTk.PhotoImage(Image.open("instructions\\error.png"))
            error_label = tk.Label(self.win_test, image=self.error_img, bg='white')
            error_label.place(relx=0.47, rely=0.60)
            self.win_test.wait_variable(self.error_ei)
            self.errors_test.append(2)
        else:
            self.errors_test.append(1)
        self.instr_now.set(1)
        self.error_ei.set(1)
        self.win_test.unbind("<KeyPress-i>")
        self.window.unbind("<KeyPress-i>")
        self.win_test.destroy()

    def i_press(self, event):
        self.end_time = time()
        self.blocks_speed.append(self.end_time - self.start_time)
        self.answers.append('i')
        self.win_test.unbind("<KeyPress-i>")
        self.window.unbind("<KeyPress-i>")
        right = self.test_error('i')
        if not right:
            self.error_img = ImageTk.PhotoImage(Image.open("instructions\\error.png"))
            error_label = tk.Label(self.win_test, image=self.error_img, bg='white')
            error_label.place(relx=0.47, rely=0.60)
            self.win_test.wait_variable(self.error_ei)
            self.errors_test.append(2)
        else:
            self.errors_test.append(1)
        self.instr_now.set(1)
        self.error_ei.set(1)
        self.win_test.unbind("<KeyPress-e>")
        self.window.unbind("<KeyPress-e>")
        self.win_test.destroy()

    def test_error(self, arg):
        right = False
        if not self.word_category:
            if self.categories_now == 11 or self.categories_now == 12:
                right = True
        elif self.word_category == 1:
            if self.categories_now == 21 or self.categories_now == 22:
                right = True
        elif self.word_category == 2:
            if self.categories_now == 11 or self.categories_now == 21:
                right = True
        else:
            if self.categories_now == 12 or self.categories_now == 22:
                right = True
        if arg == 'e':
            return not right
        return right

    def create_settings_win(self):
        check = tk.IntVar()
        self.no_error = True
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Настройки')
        win_settings.resizable(False, False)
        app_width = 375
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Исследование с помощью: ').grid(row=0, column=0)
        variableObj = tk.StringVar(self.window)
        variableObj.set(self.main_object[0])
        objectMenu = tk.OptionMenu(win_settings, variableObj, *self.main_object)
        objectMenu.grid(row=0, column=1)
        tk.Button(win_settings, text='Применить', command=lambda: check.set(1)).grid(row=1,column=1)
        win_settings.wait_variable(check)
        win_settings.destroy()
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Настройки')
        win_settings.resizable(False, False)
        app_width = 450
        app_height = 575
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        if (variableObj.get() == 'Картинки'):
            self.objects_for_study = 0
        elif (variableObj.get() == 'Слова'):
            self.objects_for_study = 1
        else:
            self.objects_for_study = 2
        var = tk.BooleanVar()
        var.set(True)
        if not self.objects_for_study: # если картинки
            self.load_pictures()
            tk.Label(win_settings, text='Количество картинок: ').grid(row=0, column=0)
            numb_of_pictures_entry =tk.Entry(win_settings)
            numb_of_pictures_entry.grid(row=0, column=1)

            tk.Checkbutton(win_settings, text='Открыть существующий Excel?', variable=var, onvalue=1,
                                          offvalue=0).grid(row=1, column=0)
            self.loadExcel = var.get()
            tk.Button(win_settings, text='Применить', command=lambda: check.set(1)).grid(row=2, column=1)
        elif self.objects_for_study == 1: # если слова
            # self.load_words()
            # tk.Label(win_settings, text='Количество слов: ').grid(row=0, column=0)
            # numb_of_words_entry = tk.Entry(win_settings)
            # numb_of_words_entry.grid(row=0, column=1)
            tk.Label(win_settings, text='Размер слова: ').grid(row=1, column=0)
            size_of_word_entry = tk.Entry(win_settings)
            size_of_word_entry.grid(row=1, column=1)
            tk.Label(win_settings, text='Цвет слов А: ').grid(row=2, column=0)
            variableColorA = tk.StringVar(win_settings)
            variableColorA.set(self.colors[0])
            objectMenuColorA = tk.OptionMenu(win_settings, variableColorA, *self.colors)
            objectMenuColorA.grid(row=2, column=1)
            tk.Label(win_settings, text='Цвет слов B: ').grid(row=3, column=0)
            variableColorB = tk.StringVar(win_settings)
            variableColorB.set(self.colors[0])
            objectMenuColorB = tk.OptionMenu(win_settings, variableColorB, *self.colors)
            objectMenuColorB.grid(row=3, column=1)
            tk.Label(win_settings, text='Категория 1 тренировка: ').grid(row=5, column=0)
            categoryA_train_entry = tk.Entry(win_settings)
            categoryA_train_entry.grid(row=5, column=1)
            tk.Label(win_settings, text='Категория 2 тренировка: ').grid(row=6, column=0)
            categoryB_train_entry = tk.Entry(win_settings)
            categoryB_train_entry.grid(row=6, column=1)
            tk.Label(win_settings, text='Атрибут 1 тренировка: ').grid(row=7, column=0)
            categoryC_train_entry = tk.Entry(win_settings)
            categoryC_train_entry.grid(row=7, column=1)
            tk.Label(win_settings, text='Атрибут 2 тренировка: ').grid(row=8, column=0)
            categoryD_train_entry = tk.Entry(win_settings)
            categoryD_train_entry.grid(row=8, column=1)
            tk.Label(win_settings, text='Категория 1 основной: ').grid(row=9, column=0)
            categoryA_main_entry = tk.Entry(win_settings)
            categoryA_main_entry.grid(row=9, column=1)
            tk.Label(win_settings, text='Категория 2 основной: ').grid(row=10, column=0)
            categoryB_main_entry = tk.Entry(win_settings)
            categoryB_main_entry.grid(row=10, column=1)
            tk.Label(win_settings, text='Атрибут 1 основной: ').grid(row=11, column=0)
            categoryC_main_entry = tk.Entry(win_settings)
            categoryC_main_entry.grid(row=11, column=1)
            tk.Label(win_settings, text='Атрибут 2 основной: ').grid(row=12, column=0)
            categoryD_main_entry = tk.Entry(win_settings)
            categoryD_main_entry.grid(row=12, column=1)
            tk.Label(win_settings, text='Слова 1 тренировка: ').grid(row=13, column=0)
            wordsA_train_entry = tk.Entry(win_settings)
            wordsA_train_entry.grid(row=13, column=1)
            tk.Label(win_settings, text='Слова 2 тренировка: ').grid(row=14, column=0)
            wordsB_train_entry = tk.Entry(win_settings)
            wordsB_train_entry.grid(row=14, column=1)
            tk.Label(win_settings, text='Слова 3 тренировка: ').grid(row=15, column=0)
            wordsC_train_entry = tk.Entry(win_settings)
            wordsC_train_entry.grid(row=15, column=1)
            tk.Label(win_settings, text='Слова 4 тренировка: ').grid(row=16, column=0)
            wordsD_train_entry = tk.Entry(win_settings)
            wordsD_train_entry.grid(row=16, column=1)
            tk.Label(win_settings, text='Слова 1 основной: ').grid(row=17, column=0)
            wordsA_main_entry = tk.Entry(win_settings)
            wordsA_main_entry.grid(row=17, column=1)
            tk.Label(win_settings, text='Слова 2 основной: ').grid(row=18, column=0)
            wordsB_main_entry = tk.Entry(win_settings)
            wordsB_main_entry.grid(row=18, column=1)
            tk.Label(win_settings, text='Слова 3 основной: ').grid(row=19, column=0)
            wordsC_main_entry = tk.Entry(win_settings)
            wordsC_main_entry.grid(row=19, column=1)
            tk.Label(win_settings, text='Слова 4 основной: ').grid(row=20, column=0)
            wordsD_main_entry = tk.Entry(win_settings)
            wordsD_main_entry.grid(row=20, column=1)
            tk.Label(win_settings, text='Посмотреть как будут выглядеть слова').grid(row=4, column=0)
            tk.Button(win_settings, text='Обновить', command=lambda: self.show_words_for_test(size_of_word_entry.get(), variableColorA.get(), variableColorB.get())).grid(row=4, column=1)
            tk.Checkbutton(win_settings, text='Открыть существующий Excel?', variable=var, onvalue=1,
                           offvalue=0).grid(row=21, column=1)
            tk.Button(win_settings, text='Применить', command=lambda: check.set(1)).grid(row=22, column=1)
            win_settings.wait_variable(check)
            self.loadExcel = var.get()
            self.size_of_font = int(size_of_word_entry.get())
            for i in range(len(self.colors)):
                if (variableColorA.get() == self.colors[i]):
                    self.indexA = i
                if (variableColorB.get() == self.colors[i]):
                    self.indexB = i
            self.categoryA_train = categoryA_train_entry.get()
            self.categoryB_train = categoryB_train_entry.get()
            self.categoryC_train = categoryC_train_entry.get()
            self.categoryD_train = categoryD_train_entry.get()
            self.categoryA_main = categoryA_main_entry.get()
            self.categoryB_main = categoryB_main_entry.get()
            self.categoryC_main = categoryC_main_entry.get()
            self.categoryD_main = categoryD_main_entry.get()
            self.wordsA_train = wordsA_train_entry.get().split(';')
            self.wordsB_train = wordsB_train_entry.get().split(';')
            self.wordsC_train = wordsC_train_entry.get().split(';')
            self.wordsD_train = wordsD_train_entry.get().split(';')
            self.wordsA_main = wordsA_main_entry.get().split(';')
            self.wordsB_main = wordsB_main_entry.get().split(';')
            self.wordsC_main = wordsC_main_entry.get().split(';')
            self.wordsD_main = wordsD_main_entry.get().split(';')

            if (len(self.wordsA_train) != len(self.wordsB_train)):
                self.error(2)
            if (len(self.wordsA_main) != len(self.wordsB_main)):
                self.error(3)
            if (len(self.wordsA_main) == 0 or len(self.wordsB_main) == 0 or len(self.wordsC_main) == 0 or len(self.wordsD_main) == 0 or
            len(self.wordsA_train) == 0 or len(self.wordsB_train) == 0 or len(self.wordsC_train) == 0 or len(self.wordsD_train) == 0):
                self.error(6)
            self.numb_for_train = len(self.wordsA_train)
            self.numb_for_main = len(self.wordsA_main)
            if self.loadExcel:
                self.ask_participant_number()
            win_settings.destroy()
        else: # если и картинки и слова
            pass
        if self.no_error:
            self.settings_changed_successfully()

    def ask_participant_number(self):
        check = tk.IntVar()
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Настройки')
        win_settings.resizable(False, False)
        app_width = 375
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Номер испытуемого: ').grid(row=0, column=0)
        number_entry = tk.Entry(win_settings)
        number_entry.grid(row=0, column=1)
        tk.Button(win_settings, text='Ок', command=lambda: check.set(1)).grid(row=1, column=1)
        win_settings.wait_variable(check)
        if number_entry.get() == '':
            temp = tk.IntVar()
            error = tk.Toplevel(self.window)
            error.wm_title('Ошибка')
            error.resizable(False, False)
            app_width = 250
            app_height = 125
            x = (self.screen_width / 2) - (app_width / 2)
            y = (self.screen_height / 2) - (app_height / 2)
            error.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
            tk.Label(error, text='Заполните все поля!').grid(row=0, column=0)
            tk.Button(error, text='Ок', command=lambda: temp.set(1)).grid(row=1, column=1)
            error.wait_variable(temp)
            win_settings.destroy()
            error.destroy()
            self.ask_participant_number()
        else:
            self.participant_number = int(number_entry.get()) - 1
            win_settings.destroy()


    def show_words_for_test(self, font_size, colorA, colorB):
        if(font_size == ''):
            font_size = 15
        self.size_of_font = int(font_size)
        for i in range(len(self.colors)):
            if(colorA == self.colors[i]):
                self.indexA = i
            if (colorB == self.colors[i]):
                self.indexB = i
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Просмотр слов')
        win_settings.resizable(False, False)
        app_width = 275
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Как будут выглядеть слова А: ').grid(row=0, column=0)
        tk.Label(win_settings, text='Пример', font=("Courier", self.size_of_font), bg='#fff', fg=self.codes_of_colors[self.indexA]).grid(row=0, column=1)
        tk.Label(win_settings, text='Как будут выглядеть слова B: ').grid(row=1, column=0)
        tk.Label(win_settings, text='Пример', font=("Courier", self.size_of_font), bg='#fff', fg=self.codes_of_colors[self.indexB]).grid(row=1, column=1)

    def settings_changed_successfully(self):
        check = tk.IntVar()
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Успешно')
        win_settings.resizable(False, False)
        app_width = 175
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Настройки изменены успешно!').grid(row=0, column=1)
        tk.Button(win_settings, text='Ок', command=lambda : check.set(1)).grid(row=1, column=1)
        win_settings.wait_variable(check)
        win_settings.destroy()

    def get_subject_info(self):
        check = tk.IntVar()
        win_settings = tk.Toplevel(self.window)
        win_settings.wm_title('Личные данные')
        win_settings.resizable(False, False)
        app_width = 300
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_settings.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_settings, text='Имя').grid(row=0, column=0)
        name_entry = tk.Entry(win_settings)
        name_entry.grid(row=0, column=1)
        tk.Label(win_settings, text='Возраст').grid(row=1, column=0)
        age_entry = tk.Entry(win_settings)
        age_entry.grid(row=1, column=1)
        tk.Label(win_settings, text='Пол').grid(row=2, column=0)
        variableObj = tk.StringVar(win_settings)
        variableObj.set(self.genders[0])
        objectMenu = tk.OptionMenu(win_settings, variableObj, *self.genders)
        objectMenu.grid(row=2, column=1)
        tk.Button(win_settings, text='Применить', command=lambda: check.set(1)).grid(row=3, column=1)
        win_settings.wait_variable(check)
        self.participant_gender = variableObj.get()
        self.participant_age = age_entry.get()
        self.participant_name = str(name_entry.get())
        if self.participant_age == '' or self.participant_name == '':
            error = tk.IntVar()
            win_error = tk.Toplevel(self.window)
            win_error.wm_title('Ошибка')
            win_error.resizable(False, False)
            app_width = 150
            app_height = 125
            x = (self.screen_width / 2) - (app_width / 2)
            y = (self.screen_height / 2) - (app_height / 2)
            win_error.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
            tk.Label(win_error, text='Введите возраст и имя!').grid(row=0, column=0)
            tk.Button(win_error, text='Ок', command=lambda: error.set(1)).grid(row=1, column=0)
            win_error.wait_variable(error)
            win_error.destroy()
            win_settings.destroy()
            self.get_subject_info()
        else:
            win_settings.destroy()
            if not self.participant_age == '' and not self.participant_name == '':
                self.get_job()

    def get_job(self):
        check = tk.IntVar()
        win_data = tk.Toplevel(self.window)
        win_data.wm_title('Работа')
        app_width = 700
        app_height = 330
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_data.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        listb = tk.Listbox(win_data, width=70, height=15)
        listb.pack()
        for i, elem in enumerate(self.works):
            listb.insert(i, elem)
        tk.Button(win_data, text='Применить',
                  command=lambda: check.set(1)).pack(pady=20)
        win_data.wait_variable(check)
        self.participant_work = listb.curselection()[0]
        win_data.destroy()
        if self.participant_work == len(self.works) - 1:
            self.get_name_job()
        if self.participant_work == 0:
            self.set_info_about_uni()

    def set_info_about_uni(self):
        check = tk.IntVar()
        win_data = tk.Toplevel(self.window)
        win_data.wm_title('Студент')
        app_width = 550
        app_height = 180
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_data.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win_data, text='Выберите свой факультет: ').grid(row=0, column=0)
        facObj = tk.StringVar(win_data)
        facObj.set(self.faculties[0])
        objectMenuFac = tk.OptionMenu(win_data, facObj, *self.faculties)
        objectMenuFac.grid(row=0, column=1)
        tk.Label(win_data, text='Выберите форму обучения: ').grid(row=1, column=0)
        formObj = tk.StringVar(win_data)
        formObj.set(self.form_of_education[0])
        objectMenuForm = tk.OptionMenu(win_data, formObj, *self.form_of_education)
        objectMenuForm.grid(row=1, column=1)
        tk.Label(win_data, text='Укажите свой курс: ').grid(row=2, column=0)
        course_entry = tk.Entry(win_data)
        course_entry.grid(row=2, column=1)
        tk.Button(win_data, text='Применить', command=lambda : check.set(1)).grid(row=3,column=1)
        win_data.wait_variable(check)
        self.participant_cour = course_entry.get()
        self.participant_fac = facObj.get()
        self.participant_form_of_edu = formObj.get()
        if self.participant_cour == '':
            error = tk.IntVar()
            temp = tk.Toplevel(self.window)
            temp.wm_title('Ошибка')
            app_width = 250
            app_height = 180
            x = (self.screen_width / 2) - (app_width / 2)
            y = (self.screen_height / 2) - (app_height / 2)
            temp.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
            tk.Label(temp, text='Заполните все поля! ').grid(row=0, column=0)
            tk.Button(temp, text='Ок', command=lambda: error.set(1)).grid(row=0, column=1)
            temp.wait_variable(error)
            temp.destroy()
            win_data.destroy()
            self.set_info_about_uni()
        else:
            win_data.destroy()

    def get_name_job(self):
        win = tk.Toplevel(self.window)
        win.wm_title('Параметры')
        app_width = 330
        app_height = 125
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Label(win, text='Профессия: ').grid(row=0, column=0)
        job_entry = tk.Entry(win)
        job_entry.grid(row=0, column=1, padx=20, pady=20)
        wait_var = tk.IntVar()
        tk.Button(win, text='Применить', command=lambda: wait_var.set(1)).grid(row=1, column=0, columnspan=2)
        win.wait_variable(wait_var)
        self.participant_work = str(job_entry.get())
        win.destroy()

    def load_pictures(self):
        pass

    # def load_words(self):
    #     try:
    #         with open('words.txt', encoding='utf-8') as file:
    #             lines = file.readlines()
    #             for line in lines:
    #                 temp = line.split(';')
    #                 self.categoryA_train = temp[0]
    #     except FileNotFoundError:
    #         self.no_error = False
    #         self.error(1)

    def clear_lists(self):
        self.block1_speed.clear()
        self.block2_speed.clear()
        self.block3_speed.clear()
        self.block4_speed.clear()
        self.block5_speed.clear()
        self.block6_speed.clear()
        self.category_order_1.clear()
        self.category_order_2.clear()
        self.category_order_3.clear()
        self.category_order_4.clear()
        self.category_order_5.clear()
        self.category_order_6.clear()
        self.errors_test.clear()
        self.bus_ans.clear()

    def error(self, type_of_error):
        check = tk.IntVar()
        self.no_error = False
        win_data = tk.Toplevel(self.window)
        win_data.wm_title('Ошибка')
        app_width = 200
        app_height = 100
        x = (self.screen_width / 2) - (app_width / 2)
        y = (self.screen_height / 2) - (app_height / 2)
        win_data.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')
        tk.Button(win_data, text='Ок', command=lambda: check.set(1)).grid(row=1, column=1)
        if type_of_error == 1:
            tk.Label(win_data, text='Не найден файл words.txt!').grid(row=0, column=0)
        elif type_of_error == 2:
            tk.Label(win_data, text='Количество слов на тренировке не совпадает!').grid(row=0, column=0)
        elif type_of_error == 3:
            tk.Label(win_data, text='Количество основных слов не совпадает!').grid(row=0, column=0)
        elif type_of_error == 4:
            tk.Label(win_data, text='Не хватает инструкций в папке!').grid(row=0, column=0)
        elif type_of_error == 5:
            tk.Label(win_data, text='Не хватает предупреждения в папке!').grid(row=0, column=0)
        elif type_of_error == 6:
            tk.Label(win_data, text='Не заполнено поле с категорией или словами!').grid(row=0, column=0)
        else:
            tk.Label(win_data, text='Что-то пошло не так, перезапустите программу!').grid(row=0, column=0)
        win_data.wait_variable(check)
        win_data.destroy()
        self.create_settings_win()

    def create_excel(self):
        error = False
        self.participant_number += 1
        if not self.loadExcel:
            wsMain = self.file1.active
            wsMain.title = "Технические условия"
            wsMain["A1"] = "Слово"
            wsMain["B1"] = "Категория (1 - категория 1; 2 - категория - 2; 3 - атрибут 1; 4 - атрибут 2"
            wsMain["C1"] = "Ответ"
            wsMain["D1"] = "Ошибка (1 - правильно, 2 - неправильно)"
            wsMain["E1"] = "Время"
            wsMain["F1"] = "Номер блока"
            wsMain["G1"] = "Порядок(1 - категория 1, атрибут - 1; 2 - категория -2, атрибут - 1)"
            wsMain["H1"] = "Имя испытуемого"
            wsMain["I1"] = "Возраст испытуемого"
            wsMain["J1"] = "Номер испытуемого"
            wsMain["K1"] = "Работа испытуемого"
            wsMain["L1"] = "Т - обучение, затем идет номер испытуемого. Пример: Т1 - обучение 1-ого испытуемого"
            wsMain["M1"] = "O - основной блок, затем идет номер блока и номер испытуемого. Пример: О2" \
                           " - основной блок для второго испытуемого, соотвественно"
        else:
            try:
                self.file1 = load_workbook('Results.xlsx')
            except FileNotFoundError:
                error = True
        ws1 = self.file1.create_sheet("T" + str(self.participant_number))
        ws2 = self.file1.create_sheet("O" + str(self.participant_number))
        row = 2
        ws1["A1"] = "Слово"
        ws1["B1"] = "Категория"
        ws1["C1"] = "Ответ"
        ws1["D1"] = "Ошибка"
        ws1["E1"] = "Время"
        ws1["F1"] = "Номер блока"
        ws1["G1"] = "Порядок"
        ws1["H1"] = "Имя"
        ws1["I1"] = "Возраст"
        ws1["J1"] = "Номер"
        ws1["K1"] = "Работа"

        ws2["A1"] = "Слово"
        ws2["B1"] = "Категория"
        ws2["C1"] = "Ответ"
        ws2["D1"] = "Ошибка"
        ws2["E1"] = "Время"
        ws2["F1"] = "Номер блока"
        ws2["G1"] = "Порядок"
        ws2["H1"] = "Имя"
        ws2["I1"] = "Возраст"
        ws2["J1"] = "Номер"
        ws2["K1"] = "Работа"
        # ws2["L1"] = "Работа"
        # ws2["M1"] = "Работа"
        # ws2["N1"] = "Работа"
        temp_var = len(self.errors_test) - 4*len(self.wordsA_main)
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_train)): # 1 блок
            if not self.category_order_1[i]:
                ws1["A" + str(row)] = str(self.wordsA_train[counter_a])
                counter_a += 1
            elif self.category_order_1[i] == 1:
                ws1["A" + str(row)] = str(self.wordsB_train[counter_b])
                counter_b += 1
            elif self.category_order_1[i] == 2:
                ws1["A" + str(row)] = str(self.wordsC_train[counter_c])
                counter_c += 1
            else:
                ws1["A" + str(row)] = str(self.wordsD_train[counter_d])
                counter_d += 1
            ws1["B" + str(row)] = str(self.category_order_1[i] + 1)
            if str(self.answers[i]) == 'e':
                ws1["C" + str(row)] = str(1)
            else:
                ws1["C" + str(row)] = str(2)
            ws1["D" + str(row)] = str(self.errors_test[i])
            ws1["E" + str(row)] = str(self.blocks_speed[i])
            ws1["F" + str(row)] = '1'
            ws1["G" + str(row)] = str(self.categories_order[0])
            ws1["H" + str(row)] = str(self.participant_name)
            ws1["I" + str(row)] = str(self.participant_age)
            ws1["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws1["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws1["K" + str(row)] = str(self.works[self.participant_work])
                ws1["L" + str(row)] = str(self.participant_cour)
                ws1["M" + str(row)] = str(self.participant_fac)
                ws1["N" + str(row)] = str(self.participant_form_of_edu)
                ws1["M1"] = "Факультет"
                ws1["L1"] = "Курс"
                ws1["N1"] = "Форма обучения"
            else:
                ws1["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        temp = row
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_train)): # 2 блок
            if not self.category_order_2[i]:
                ws1["A" + str(row)] = str(self.wordsA_train[counter_a])
                counter_a += 1
            elif self.category_order_2[i] == 1:
                ws1["A" + str(row)] = str(self.wordsB_train[counter_b])
                counter_b += 1
            elif self.category_order_2[i] == 2:
                ws1["A" + str(row)] = str(self.wordsC_train[counter_c])
                counter_c += 1
            else:
                ws1["A" + str(row)] = str(self.wordsD_train[counter_d])
                counter_d += 1
            ws1["B" + str(row)] = str(self.category_order_2[i] + 1)
            ws1["F" + str(row)] = '2'
            ws1["G" + str(row)] = str(self.categories_order[1])
            ws1["H" + str(row)] = str(self.participant_name)
            ws1["I" + str(row)] = str(self.participant_age)
            ws1["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws1["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws1["K" + str(row)] = str(self.works[self.participant_work])
                ws1["L" + str(row)] = str(self.participant_cour)
                ws1["M" + str(row)] = str(self.participant_fac)
                ws1["N" + str(row)] = str(self.participant_form_of_edu)
                ws1["M1"] = "Факультет"
                ws1["L1"] = "Курс"
                ws1["N1"] = "Форма обучения"
            else:
                ws1["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        for i in range(4*len(self.wordsA_train), 2*4*len(self.wordsA_train)):
            ws1["E" + str(temp)] = str(self.blocks_speed[i])
            if str(self.answers[i]) == 'e':
                ws1["C" + str(temp)] = str(1)
            else:
                ws1["C" + str(temp)] = str(2)
            ws1["D" + str(temp)] = str(self.errors_test[i])
            temp += 1
        ws1["O1"] = "Ответы"
        row = 2
        for i in range(len(self.bus_ans)):
            ws1["O" + str(row)] = self.bus_ans[i]
            row += 1
        temp = row = 2
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_main)): # 3 блок
            if not self.category_order_3[i]:
                ws2["A" + str(row)] = str(self.wordsA_main[counter_a])
                counter_a += 1
            elif self.category_order_3[i] == 1:
                ws2["A" + str(row)] = str(self.wordsB_main[counter_b])
                counter_b += 1
            elif self.category_order_3[i] == 2:
                ws2["A" + str(row)] = str(self.wordsC_main[counter_c])
                counter_c += 1
            else:
                ws2["A" + str(row)] = str(self.wordsD_main[counter_d])
                counter_d += 1
            ws2["B" + str(row)] = str(self.category_order_3[i] + 1)
            ws2["F" + str(row)] = '3'
            ws2["G" + str(row)] = str(self.categories_order[2])
            ws2["H" + str(row)] = str(self.participant_name)
            ws2["I" + str(row)] = str(self.participant_age)
            ws2["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws2["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
                ws2["L" + str(row)] = str(self.participant_cour)
                ws2["M" + str(row)] = str(self.participant_fac)
                ws2["N" + str(row)] = str(self.participant_form_of_edu)
                ws2["M1"] = "Факультет"
                ws2["L1"] = "Курс"
                ws2["N1"] = "Форма обучения"
            else:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        for i in range(2*4*len(self.wordsA_train), 2*4*len(self.wordsA_train) + 4*len(self.wordsA_main)):
            ws2["E" + str(temp)] = str(self.blocks_speed[i])
            if str(self.answers[i]) == 'e':
                ws2["C" + str(temp)] = str(1)
            else:
                ws2["C" + str(temp)] = str(2)
            ws2["D" + str(temp)] = str(self.errors_test[i])
            temp += 1
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_main)): # 4 блок
            if not self.category_order_4[i]:
                ws2["A" + str(row)] = str(self.wordsA_main[counter_a])
                counter_a += 1
            elif self.category_order_4[i] == 1:
                ws2["A" + str(row)] = str(self.wordsB_main[counter_b])
                counter_b += 1
            elif self.category_order_4[i] == 2:
                ws2["A" + str(row)] = str(self.wordsC_main[counter_c])
                counter_c += 1
            else:
                ws2["A" + str(row)] = str(self.wordsD_main[counter_d])
                counter_d += 1
            ws2["B" + str(row)] = str(self.category_order_4[i] + 1)
            ws2["F" + str(row)] = '4'
            ws2["G" + str(row)] = str(self.categories_order[3])
            ws2["H" + str(row)] = str(self.participant_name)
            ws2["I" + str(row)] = str(self.participant_age)
            ws2["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws2["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
                ws2["L" + str(row)] = str(self.participant_cour)
                ws2["M" + str(row)] = str(self.participant_fac)
                ws2["N" + str(row)] = str(self.participant_form_of_edu)
                ws2["M1"] = "Факультет"
                ws2["L1"] = "Курс"
                ws2["N1"] = "Форма обучения"
            else:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        for i in range(2*4*len(self.wordsA_train) + 4*len(self.wordsA_main), 2*4*len(self.wordsA_train) + 2*4*len(self.wordsA_main)):
            ws2["E" + str(temp)] = str(self.blocks_speed[i])
            if str(self.answers[i]) == 'e':
                ws2["C" + str(temp)] = str(1)
            else:
                ws2["C" + str(temp)] = str(2)
            ws2["D" + str(temp)] = str(self.errors_test[i])
            temp += 1
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_main)): # 5 блок
            if not self.category_order_5[i]:
                ws2["A" + str(row)] = str(self.wordsA_main[counter_a])
                counter_a += 1
            elif self.category_order_5[i] == 1:
                ws2["A" + str(row)] = str(self.wordsB_main[counter_b])
                counter_b += 1
            elif self.category_order_5[i] == 2:
                ws2["A" + str(row)] = str(self.wordsC_main[counter_c])
                counter_c += 1
            else:
                ws2["A" + str(row)] = str(self.wordsD_main[counter_d])
                counter_d += 1
            ws2["B" + str(row)] = str(self.category_order_5[i] + 1)
            ws2["F" + str(row)] = '5'
            ws2["G" + str(row)] = str(self.categories_order[4])
            ws2["H" + str(row)] = str(self.participant_name)
            ws2["I" + str(row)] = str(self.participant_age)
            ws2["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws2["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
                ws2["L" + str(row)] = str(self.participant_cour)
                ws2["M" + str(row)] = str(self.participant_fac)
                ws2["N" + str(row)] = str(self.participant_form_of_edu)
                ws2["M1"] = "Факультет"
                ws2["L1"] = "Курс"
                ws2["N1"] = "Форма обучения"
            else:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        for i in range(2*4*len(self.wordsA_train) + 2*4*len(self.wordsA_main), 2*4*len(self.wordsA_train) + 3*4*len(self.wordsA_main)):
            ws2["E" + str(temp)] = str(self.blocks_speed[i])
            if str(self.answers[i]) == 'e':
                ws2["C" + str(temp)] = str(1)
            else:
                ws2["C" + str(temp)] = str(2)
            ws2["D" + str(temp)] = str(self.errors_test[i])
            temp += 1
        counter_a = counter_b = counter_c = counter_d = 0
        for i in range(4*len(self.wordsA_main)): # 6 блок
            if not self.category_order_6[i]:
                ws2["A" + str(row)] = str(self.wordsA_main[counter_a])
                counter_a += 1
            elif self.category_order_6[i] == 1:
                ws2["A" + str(row)] = str(self.wordsB_main[counter_b])
                counter_b += 1
            elif self.category_order_6[i] == 2:
                ws2["A" + str(row)] = str(self.wordsC_main[counter_c])
                counter_c += 1
            else:
                ws2["A" + str(row)] = str(self.wordsD_main[counter_d])
                counter_d += 1
            ws2["B" + str(row)] = str(self.category_order_6[i] + 1)
            ws2["F" + str(row)] = '6'
            ws2["G" + str(row)] = str(self.categories_order[5])
            ws2["H" + str(row)] = str(self.participant_name)
            ws2["I" + str(row)] = str(self.participant_age)
            ws2["J" + str(row)] = str(self.participant_number)
            if self.participant_work == len(self.works) - 1:
                ws2["K" + str(row)] = str(self.job)
            elif not self.participant_work:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
                ws2["L" + str(row)] = str(self.participant_cour)
                ws2["M" + str(row)] = str(self.participant_fac)
                ws2["N" + str(row)] = str(self.participant_form_of_edu)
                ws2["M1"] = "Факультет"
                ws2["L1"] = "Курс"
                ws2["N1"] = "Форма обучения"
            else:
                ws2["K" + str(row)] = str(self.works[self.participant_work])
            row += 1
        for i in range(2*4*len(self.wordsA_train) + 3*4*len(self.wordsA_main), 2*4*len(self.wordsA_train) + 4*4*len(self.wordsA_main)):
            ws2["E" + str(temp)] = str(self.blocks_speed[i])
            if str(self.answers[i]) == 'e':
                ws2["C" + str(temp)] = str(1)
            else:
                ws2["C" + str(temp)] = str(2)
            ws2["D" + str(temp)] = str(self.errors_test[i])
            temp += 1
        if error:
            error_win = tk.Toplevel(self.window)
            error_win.wm_title('Ошибка')
            error_win.geometry("650x100")
            tk.Label(error_win, text='Файл с результами не найден. Данные сохранены '
                                     'в ' + "ErrorResults" + str(self.participant_number) + ".xlsx").grid(row=0,
                                                                                                          column=1)
            self.file1.save("ErrorResults" + str(self.participant_number) + ".xlsx")
        else:
            try:
                self.file1.save("Results.xlsx")
            except PermissionError:
                error_win = tk.Toplevel(self.window)
                error_win.wm_title('Ошибка')
                error_win.geometry("650x100")
                tk.Label(error_win, text='Файл с результатами открыт, закройте его. Данные сохранены '
                                         'в ' + "ErrorResults" + str(self.participant_number) + ".xlsx").grid(row=0,
                                                                                                              column=1)
                self.file1.save("ErrorResults" + str(self.participant_number) + ".xlsx")


program = MyProgram()
program.start()